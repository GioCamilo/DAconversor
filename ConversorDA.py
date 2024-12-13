import sys
import json
from PySide6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QFileDialog, QWidget, QLabel, QLineEdit, QHBoxLayout, QFrame, QListWidget, QHeaderView, QMessageBox, QProgressBar)
from PySide6.QtCore import Qt, QTimer
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook
from PySide6.QtWidgets import QMessageBox
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook


HEADERS = [
    "Discriminação das despesas não tributável", "Pagas pela Comissária", "NF FILHA",
    "NF MÃE", "Nº DOC", "CFOP", "CÓD PARCEIRO", "PROCESSO", "CONTA CONTÁBIL",
    "VALOR LIQUIDO", "ICMS %", "VALOR BRUTO", "CLASSIFICAÇÃO", "PIS",
    "CONFINS", "ICMS", "PEDIDO", "ITEM" , "COM EMISSÃO DE NFe"
]

class PDFTableExtractorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Conversor Despesa Aduaneira v2025")
        self.resize(1200, 800)

        # Carregar os caminhos dos arquivos do JSON
        self.load_paths_from_json(r"Q:\Fiscal\PROGRAMAS FISCAIS\PYTHON SOLUTIONS\Conversor DA\CONFIG\config\config.json")
        
        # Widget central e layout principal
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)

        # Criando o layout horizontal principal
        self.fields_layout = QHBoxLayout()
        self.fields_layout.setSpacing(10)  # Espaço entre os campos
        self.main_layout.addLayout(self.fields_layout)

        # Definindo os campos de entrada
        self.reference_label = QLabel("Referência:")
        self.reference_input = QLineEdit()
        self.importer_label = QLabel("Importador:")
        self.importer_input = QLineEdit()
        self.total_label = QLabel("Total não Trib. (-):")
        self.total_input = QLineEdit()
        # Adicionar mais três campos vazios
        self.check_1_label = QLabel("Total não Trib(sem frete):")
        self.field_check_1 = QLineEdit()
        self.field_check_1.setReadOnly(True)
        self.check_2_label = QLabel("Total Liquido:")
        self.field_check_2 = QLineEdit()
        self.field_check_2.setReadOnly(True)
        self.check_3_label = QLabel("Check:")
        self.field_check_3 = QLineEdit()
        self.field_check_3.setReadOnly(True)
        self.total_bruto_label = QLabel("Total Bruto:")
        self.total_bruto_input = QLineEdit()
        self.total_cd_label = QLabel("Total bruto CD:")
        self.total_cd_input = QLineEdit()
        self.total_cabruto_label = QLabel("Total bruto CA:")
        self.total_cabruto_input = QLineEdit()
        self.total_caliqui_label = QLabel("Total liquido CA:")
        self.total_caliqui_input = QLineEdit()

        # Criando os layouts verticais para cada coluna
        coluna1 = QVBoxLayout()
        coluna2 = QVBoxLayout()
        coluna3 = QVBoxLayout()
        coluna4 = QVBoxLayout()
        coluna5 = QVBoxLayout()

        # Adicionando os campos na coluna 1
        coluna1.addWidget(self.reference_label)
        coluna1.addWidget(self.reference_input)
        coluna1.addWidget(self.check_1_label)
        coluna1.addWidget(self.field_check_1)

        # coluna2
        coluna2.addWidget(self.importer_label)
        coluna2.addWidget(self.importer_input)
        coluna2.addWidget(self.check_2_label)
        coluna2.addWidget(self.field_check_2)

        # coluna3
        coluna3.addWidget(self.total_label)
        coluna3.addWidget(self.total_input)
        coluna3.addWidget(self.check_3_label)
        coluna3.addWidget(self.field_check_3)
        
        #coluna4
        coluna4.addWidget(self.total_bruto_label)
        coluna4.addWidget(self.total_bruto_input)
        coluna4.addWidget(self.total_cd_label)
        coluna4.addWidget(self.total_cd_input)
        
        #coluna5
        coluna5.addWidget(self.total_cabruto_label)
        coluna5.addWidget(self.total_cabruto_input)
        coluna5.addWidget(self.total_caliqui_label)
        coluna5.addWidget(self.total_caliqui_input)
        # adicionando coluna no layout principal
        self.fields_layout.addLayout(coluna1)
        self.fields_layout.addLayout(coluna2)  # Corrigido: adicionando coluna2
        self.fields_layout.addLayout(coluna3)
        self.fields_layout.addLayout(coluna4)
        self.fields_layout.addLayout(coluna5)
        
        
        # Área superior com botões
        self.top_layout = QHBoxLayout()
        self.main_layout.addLayout(self.top_layout)

        # Área superior com botões
        self.top_layout = QHBoxLayout()
        self.main_layout.addLayout(self.top_layout)  # Adiciona o top_layout no layout principal

        self.select_button = QPushButton("Selecionar PDFs")
        self.select_button.clicked.connect(self.select_pdfs)
        self.top_layout.addWidget(self.select_button)

        self.export_button = QPushButton("Exportar para Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.top_layout.addWidget(self.export_button)

        # Adiciona um espaçador flexível para empurrar os botões para a esquerda
        self.top_layout.addStretch()

        # Redimensionar botões
        self.select_button.setFixedSize(120, 30)  # Define largura e altura para o botão 'Selecionar PDFs'
        self.export_button.setFixedSize(120, 30)  # Define largura e altura para o botão 'Exportar para Excel'

        
        # Layout principal com lista de arquivos e tabela
        self.main_frame = QHBoxLayout()
        self.main_frame.setSpacing(20)  # Espaço entre a lista e a tabela
        self.main_layout.addLayout(self.main_frame)

        # Lista de arquivos
        self.file_list = QListWidget()
        self.file_list.itemClicked.connect(self.on_file_select)
        self.main_frame.addWidget(self.file_list, 1)

        # Tabela
        self.table = QTableWidget()
        self.table.setColumnCount(len(HEADERS))
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.horizontalHeader().setStretchLastSection(True)  # Estica a última coluna
        self.main_frame.addWidget(self.table, 3)

        # Configuração dos cabeçalhos e larguras das colunas
        column_widths = [
            250, 150, 120, 120, 100, 100, 150, 120, 150, 120, 
            80, 100, 200, 80, 80, 80, 80, 80, 200
        ]

        if len(HEADERS) != len(column_widths):
            raise ValueError("O número de cabeçalhos e larguras de coluna deve ser igual.")

        for index, header in enumerate(HEADERS):
            self.table.setColumnWidth(index, column_widths[index])
            self.table.horizontalHeader().setSectionResizeMode(index, QHeaderView.Interactive)

        # Conectar o evento de edição 
        self.table.cellChanged.connect(self.save_table_changes)
        
        # Barra de progresso
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.main_layout.addWidget(self.progress_bar)
        
        # Dados internos
        self.loaded_files = {}
        self.tabelas_importadas = {}

        # Timer para carregamento gradual
        self.load_timer = QTimer()
        self.load_timer.timeout.connect(self.load_table_data)
        self.lazy_load_index = 0
        
    def load_paths_from_json(self, json_path):
        try:
            # Carregar o arquivo JSON
            with open(json_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                # Ajustar os caminhos carregados
                self.banco_de_dados = data.get("banco_de_dados")
                self.controle_path = data.get("controle_path")

                # Imprimir os caminhos carregados para diagnóstico
                print(f"Caminho do Banco de Dados: {self.banco_de_dados}")
                print(f"Caminho da Planilha de Controle: {self.controle_path}")

                # Verificar se os caminhos foram carregados corretamente
                if not self.banco_de_dados:
                    print("Erro: Caminho do banco de dados não encontrado no JSON.")
                if not self.controle_path:
                    print("Erro: Caminho da planilha de controle não encontrado no JSON.")

        except FileNotFoundError:
            print(f"Erro: Arquivo JSON '{json_path}' não encontrado.")
        except json.JSONDecodeError as e:
            print(f"Erro ao decodificar o JSON: {e}")
        except Exception as e:
            print(f"Erro ao carregar o arquivo JSON: {e}")

        # Caso haja falha ao carregar os caminhos, definir como None
        if not hasattr(self, 'banco_de_dados'):
            self.banco_de_dados = None
        if not hasattr(self, 'controle_path'):
            self.controle_path = None

    def select_pdfs(self):
        # Abre um diálogo para selecionar arquivos PDF
        file_paths, _ = QFileDialog.getOpenFileNames(self, "Selecionar PDFs", "", "PDF Files (*.pdf)")
        if file_paths:
            self.loaded_files = {f"PDF {i + 1}": path for i, path in enumerate(file_paths)}

            # Atualizar a lista de arquivos
            self.file_list.clear()
            for name in self.loaded_files.keys():
                self.file_list.addItem(name)

            # Configurar a barra de progresso
            self.progress_bar.setMaximum(len(file_paths))
            self.progress_bar.setValue(0)

            # Processar e criar as tabelas automaticamente após importar os PDFs
            for idx, file_path in enumerate(self.loaded_files.values(), start=1):
                try:
                    referencia, importador, total_nao_trib = self.extract_values_from_pdf(file_path)

                    # Atualizar os campos de entrada
                    self.reference_input.setText(referencia)
                    self.importer_input.setText(importador)
                    self.total_input.setText(total_nao_trib)

                    # Extrair dados do PDF
                    data = self.extract_columns_from_pdf(file_path)

                    # Buscar dados adicionais no Excel
                    excel_data = None
                    if referencia:
                        excel_data = self.search_in_excel(referencia)

                    # Atualizar a tabela com os dados processados
                    self.update_table(data, excel_data)
                    self.compare_and_fill_classification()
                    self.calculate_check()
                    self.calcular_total_bruto()
                    self.calcular_total_bruto_cd()
                    self.calcular_total_bruto_ca()
                    self.calcular_total_liquido_ca()

                    # Armazenar os dados extraídos na estrutura 'tabelas_importadas'
                    if file_path not in self.tabelas_importadas:
                        table_data = []
                        for row_idx in range(self.table.rowCount()):
                            row = []
                            for col_idx in range(self.table.columnCount()):
                                item = self.table.item(row_idx, col_idx)
                                row.append(item.text() if item else "--")
                            table_data.append(row)

                        # Armazenar os dados como um dicionário, não apenas a lista da tabela
                        self.tabelas_importadas[file_path] = {
                            "referencia": referencia,
                            "importador": importador,
                            "total_nao_trib": total_nao_trib,
                            "data": table_data
                        }

                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao processar o PDF: {e}")
                finally:
                    # Atualizar a barra de progresso
                    self.progress_bar.setValue(idx)
                print(f"Dados armazenados para {file_path}: {self.tabelas_importadas[file_path]}")

    def on_file_select(self, item):
        """
        Atualiza a tabela, os campos e recalcula os valores com base no PDF selecionado.
        """
        file_path = self.loaded_files.get(item.text())
        if file_path:
            try:
                # Recuperar os dados armazenados
                table_data = self.tabelas_importadas[file_path]

                # Atualizar os campos de entrada
                self.reference_input.setText(table_data.get("referencia", ""))
                self.importer_input.setText(table_data.get("importador", ""))
                self.total_input.setText(table_data.get("total_nao_trib", ""))

                # Atualizar a tabela
                tabela_dados = table_data.get("data", [])
                self.table.blockSignals(True)  # Evita disparar eventos desnecessários
                self.table.clearContents()
                self.table.setRowCount(len(tabela_dados))
                for row_idx, row in enumerate(tabela_dados):
                    for col_idx, value in enumerate(row):
                        self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
                self.table.blockSignals(False)

                # Recalcular os valores de Check
                self.calculate_check()
                self.calcular_total_bruto()
                self.calcular_total_bruto_cd()
                self.calcular_total_bruto_ca()
                self.calcular_total_liquido_ca()
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao processar o PDF: {e}")




    def on_table_data_changed(self):
        """
        Atualiza os campos de check sempre que os dados da tabela forem alterados pelo usuário.
        """
        self.calculate_check()

    def load_table_data(self):
        """
        Carrega os dados da tabela de forma gradual para melhorar a experiência do usuário.
        """
        if self.lazy_load_index < len(self.current_table_data):
            row_data = self.current_table_data[self.lazy_load_index]
            for col_idx, value in enumerate(row_data):
                self.table.setItem(self.lazy_load_index, col_idx, QTableWidgetItem(str(value)))
            self.lazy_load_index += 1
        else:
            self.load_timer.stop()

    def extract_values_from_pdf(self, file_path):
        """
        Extrai valores gerais do PDF, como referência, importador e total não tributável.   
        Garante que apenas o código de referência seja retornado.
        """
        referencia, importador, total_nao_trib = "", "", ""
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        lines = text.split("\n")
                        for line in lines:
                            # Localizar e extrair apenas o código da referência
                            if "Referência:" in line:
                                referencia = line.split("Referência:")[1].strip().split()[0]  # Extrai apenas o código
                            elif "Importador/Exportador:" in line:
                                importador = line.split(":")[1].strip()
                            elif "Total não Trib. (-):" in line:
                                total_nao_trib = line.split(":")[-1].strip()
            return referencia, importador, total_nao_trib
        except Exception as e:
            raise ValueError(f"Erro ao processar PDF: {e}")

    def extract_columns_from_pdf(self, file_path):
        """Extrai informações das colunas 'Discriminação' e 'Pagas pela Comissária'."""
        data = []
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        lines = text.split("\n")
                        start_index = 0
                        for i, line in enumerate(lines):
                            if "Discriminação das despesas não tributável" in line and "Pagas pela Comissária" in line:
                                start_index = i + 1
                                break
                        for line in lines[start_index:]:
                            if not line.strip() or "total" in line.lower():
                                break
                            parts = line.split()
                            column_1 = " ".join(parts[:-1])
                            column_2 = parts[-1] if parts else ""
                            row_data = [column_1, column_2] + ["--"] * (len(HEADERS) - 2)
                            data.append(row_data)
            return data
        except Exception as e:
            raise ValueError(f"Erro ao processar as colunas do PDF: {e}")

    def search_in_excel(self, referencia):
        """Busca informações adicionais no Excel com base na referência."""
        if not self.controle_path:
            QMessageBox.critical(self, "Erro", "Caminho da planilha de controle não foi carregado corretamente.")
            return None

        try:
            # Carregar o workbook
            workbook = load_workbook(self.controle_path)
            sheet = workbook.active  # Selecionar a aba ativa

            # Iterar pelas linhas da planilha
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Ignorar cabeçalho
                print(f"Lendo linha {row_idx}: {row}")  # Diagnóstico: Mostrar os valores da linha

                # Verificar se o valor da coluna AD corresponde à referência
                if str(row[29]).strip() == referencia.strip():  # Normalizar espaços e verificar a correspondência
                    print(f"Referência encontrada na linha {row_idx}: {row[29]}")  # Diagnóstico: Mostrar quando encontrada
                    return {
                        "coluna_g": row[6],   # Coluna G
                        "coluna_f": row[5],   # Coluna F
                        "coluna_m": row[12],  # Coluna M
                        "coluna_k": row[10],  # Coluna K
                        "coluna_t": row[19],  # Coluna T
                        "coluna_u": row[20],  # Coluna U
                        "coluna_x": row[23],  # Coluna X
                        "coluna_aa": row[26], # Coluna AA
                        "coluna_ad": row[29], # Coluna AD
                    }

            # Caso a referência não seja encontrada
            print(f"Referência '{referencia}' não encontrada no Excel.")
            return None  

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao acessar o Excel: {e}")
            return None

    def update_table(self, data, excel_data=None, coluna_b_values=None):
        """
        Atualiza a tabela com os dados do PDF, aplica cálculos específicos para PIS, CONFINS, e ICMS.
        """
        self.table.setRowCount(len(data))  # Define o número de linhas da tabela

        for row_idx, row_data in enumerate(data):
            # Garantir que o número de colunas corresponda ao número de HEADERS
            row_data = row_data[:len(HEADERS)] + ["--"] * (len(HEADERS) - len(row_data))

            # Atualizar lógica específica para colunas
            row_data[9] = row_data[1]  # Atualizar a coluna 10 (índice 9)

            # Cálculo para PIS, CONFINS
            try:
                # Limpar e converter o valor da Coluna 2
                valor_liquido = row_data[1].replace(".", "").replace(",", ".").strip()
                valor_liquido = float(valor_liquido) if valor_liquido else 0.0

                # Percentual de ICMS, PIS e CONFINS
                percentual_icms = 0.18  # ICMS 18%
                percentual_pis = 0.0165  # PIS 1.65%
                percentual_confins = 0.076  # CONFINS 7.6%

                # Calcular PIS
                valor_pis = valor_liquido * percentual_pis
                row_data[13] = f"{valor_pis:,.2f}".replace(".", ",")  # Coluna PIS (índice 13)

                # Calcular CONFINS
                valor_confins = valor_liquido * percentual_confins
                row_data[14] = f"{valor_confins:,.2f}".replace(".", ",")  # Coluna CONFINS (índice 14)

            except ValueError:
                row_data[15] = "--"  # ICMS
                row_data[13] = "--"  # PIS
                row_data[14] = "--"  # CONFINS

            # Atualizar a lógica para "VALOR BRUTO" (Coluna 12)
            try:
                # Calcular Valor Bruto com a fórmula
                valor_bruto = valor_liquido / (1 - percentual_icms)
                valor_bruto_arredondado = round(valor_bruto, 2)  # Arredonda para 2 casas decimais

                # Formatar o valor com separador de milhar e decimal
                valor_bruto_formatado = f"{valor_bruto_arredondado:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")  

                row_data[11] = valor_bruto_formatado  # Atualiza a Coluna 12
                
            # Calcular ICMS
                valor_icms = valor_bruto * percentual_icms
                row_data[15] = f"{valor_icms:,.2f}".replace(".", ",")  # Coluna ICMS (índice 15)
                
            except ZeroDivisionError:
                row_data[11] = "--"

            # Atualizar as colunas com dados do Excel, se disponíveis
            if excel_data:
                row_data[3] = excel_data.get("coluna_g", "--")  # Coluna 4: G
                row_data[4] = excel_data.get("coluna_f", "--")  # Coluna 5: F
                row_data[5] = excel_data.get("coluna_m", "--")  # Coluna 6: M
                row_data[6] = excel_data.get("coluna_k", "--")  # Coluna 7: K
                row_data[7] = excel_data.get("coluna_ad", "--")  # Coluna 8: AD
                row_data[8] = excel_data.get("coluna_x", "--")  # Coluna 9: X
                row_data[10] = excel_data.get("coluna_aa", "--")  # Coluna 10: AA
                row_data[16] = excel_data.get("coluna_t", "--")  # Coluna 17: T
                row_data[17] = excel_data.get("coluna_u", "--")  # Coluna 18: U
                row_data[2] = ""   # Coluna 3: B        # Coluna 3: B

            
            # Atualizar lógica de outras colunas (se necessário)
            valor_col3 = row_data[2] if len(row_data) > 2 else ""  # Garantir que a coluna 3 existe
            if isinstance(valor_col3, str) and valor_col3.strip() in ["", "--"]:
                row_data[18] = "SEM EMISSÃO DE NFe"  # Coluna 19 (índice 18)
            else:
                row_data[18] = "COM EMISSÃO DE NFe"  # Coluna 19 (índice 18)

            # Aplicar regra para CFOP (Coluna 6, índice 5)
            descricao_coluna1 = row_data[0].strip().lower() if row_data[0] else ""
            cfop_valor = row_data[5]

            if "armazenagem" in descricao_coluna1 and isinstance(cfop_valor, str):
                # Se a descrição contiver "armazenagem", alterar para "CA"
                partes_cfop = cfop_valor.split("/")
                if len(partes_cfop) == 2:
                    partes_cfop[1] = "CA"
                    row_data[5] = "/".join(partes_cfop)
            elif isinstance(cfop_valor, str):
                # Para demais despesas, alterar para "CD"
                partes_cfop = cfop_valor.split("/")
                if len(partes_cfop) == 2:
                    partes_cfop[1] = "CD"
                    row_data[5] = "/".join(partes_cfop)

            # Inserir os dados na tabela
            for col_idx, value in enumerate(row_data):
                self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))



    

    def compare_and_fill_classification(self):
        """
        Atualiza a Coluna 13 (CLASSIFICAÇÃO) da tabela com base na Coluna B do arquivo Excel.
        """
        # Utilizar o caminho do JSON carregado
        if not self.banco_de_dados:
            QMessageBox.critical(self, "Erro", "Caminho do banco de dados não foi carregado corretamente.")
            return

        try:
            # Carregar o arquivo Excel
            workbook = load_workbook(self.banco_de_dados)
            sheet = workbook.active  # Seleciona a planilha ativa

            # Criar um dicionário para mapear as descrições da Coluna A com os valores da Coluna B
            despesas_map = {}
            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):  # Coluna A (descrição) e B (valor)
                descricao = row[0]
                valor = row[1]

                if descricao is not None:  # Apenas normalizar a descrição
                    descricao_normalizada = str(descricao).strip().lower()
                    despesas_map[descricao_normalizada] = valor if valor is not None else "--"

            # Iterar pela tabela do aplicativo
            for row_idx in range(self.table.rowCount()):
                descricao_tabela = self.table.item(row_idx, 0)

                if descricao_tabela:
                    descricao_tabela_texto = descricao_tabela.text()
                    descricao_tabela_normalizada = str(descricao_tabela_texto).strip().lower()

                    print(f"Descrição da Tabela (linha {row_idx}): {descricao_tabela_normalizada}")

                    # Verificar se a descrição está no mapeamento do Excel
                    if descricao_tabela_normalizada in despesas_map:
                        valor = despesas_map[descricao_tabela_normalizada]
                        valor_formatado = (
                            f"{valor:,.2f}".replace(".", ",") if isinstance(valor, (int, float)) else valor
                        )
                        print(f"Atualizando valor para a linha {row_idx}, coluna CLASSIFICAÇÃO: {valor_formatado}")
                        self.table.setItem(row_idx, 12, QTableWidgetItem(valor_formatado))  # Índice 12 -> Coluna 13
                    else:
                        print(f"Descrição não encontrada no mapeamento: {descricao_tabela_normalizada}")
                        self.table.setItem(row_idx, 12, QTableWidgetItem("--"))
                else:
                    print(f"Descrição vazia na linha {row_idx}")
                    self.table.setItem(row_idx, 12, QTableWidgetItem("--"))

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao acessar o arquivo Excel: {e}")


    def on_file_select(self, item):
        """
        Atualiza a tabela, os campos e recalcula os valores com base no PDF selecionado.
        """
        file_path = self.loaded_files.get(item.text())
        if file_path:
            try:
                # Verificar se os dados do PDF já estão armazenados em 'tabelas_importadas'
                if file_path in self.tabelas_importadas:
                    table_data = self.tabelas_importadas[file_path]

                    # Atualizar os campos de entrada com os valores correspondentes
                    referencia = table_data.get("referencia", "")
                    importador = table_data.get("importador", "")
                    total_nao_trib = table_data.get("total_nao_trib", "")

                    self.reference_input.setText(referencia)
                    self.importer_input.setText(importador)
                    self.total_input.setText(total_nao_trib)

                    # Atualizar a tabela
                    tabela_dados = table_data.get("data", [])
                    self.table.blockSignals(True)  # Evita disparar eventos desnecessários
                    self.table.clearContents()
                    self.table.setRowCount(len(tabela_dados))
                    for row_idx, row in enumerate(tabela_dados):
                        for col_idx, value in enumerate(row):
                            self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
                    self.table.blockSignals(False)

                    # Recalcular os valores de Check
                    self.calculate_check()
                    self.calcular_total_bruto()
                    self.calcular_total_bruto_ca()
                    self.calcular_total_bruto_cd()
                    self.calcular_total_liquido_ca()
                else:
                    QMessageBox.warning(self, "Aviso", f"Dados do PDF '{item.text()}' não encontrados.")

            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao processar o PDF: {e}")


    def save_table_changes(self, row, column):
        """
        Salva as alterações feitas pelo usuário na tabela atual em 'tabelas_importadas'.
        Atualiza automaticamente a coluna 19 se a coluna 3 for editada.
        """
        # Bloquear sinais
        self.table.blockSignals(True)
        # Formatar o novo valor da célula
        if column == 11:  # Coluna 12
            new_value = self.table.item(row, column).text()
            if new_value not in ["", "--"]:
                new_value = new_value.replace(".", "").replace(",", ".")
                self.table.setItem(row, column, QTableWidgetItem(new_value))

        current_item = self.file_list.currentItem()
        if current_item:
            file_path = self.loaded_files.get(current_item.text())
            if file_path and file_path in self.tabelas_importadas:
                # Obter o novo valor da célula editada
                new_value = self.table.item(row, column).text() if self.table.item(row, column) else "--"

                # Atualizar o valor na estrutura armazenada
                if "data" in self.tabelas_importadas[file_path]:
                    self.tabelas_importadas[file_path]["data"][row][column] = new_value

                # Atualizar a coluna 19 (índice 18) se a coluna 3 (índice 2) foi editada
                if column == 2:  # Índice 2 corresponde à coluna 3
                    if new_value.strip():  # Se a coluna 3 tiver algum valor
                        self.table.blockSignals(True)  # Bloquear sinais para evitar loop
                        self.table.setItem(row, 18, QTableWidgetItem("COM EMISSÃO DE NFe"))
                        self.tabelas_importadas[file_path]["data"][row][18] = "COM EMISSÃO DE NFe"
                        self.table.blockSignals(False)
                    else:  # Se o valor estiver vazio
                        self.table.blockSignals(True)
                        self.table.setItem(row, 18, QTableWidgetItem("SEM EMISSÃO DE NFe"))
                        self.tabelas_importadas[file_path]["data"][row][18] = "SEM EMISSÃO DE NFe"
                        self.table.blockSignals(False)
                        self.calcular_total_bruto()
                        self.calcular_total_bruto_cd()

    def calculate_check(self):
        try:
            # Obter o valor de Total não Trib. (-)
            total_nao_trib = self.total_input.text().strip()
            if not total_nao_trib:
                raise ValueError("O campo 'Total não Trib. (-)' está vazio.")
            
            # Limpar o formato numérico
            total_nao_trib = float(total_nao_trib.replace(".", "").replace(",", "."))

            # Inicializar valores para cálculo
            total_valor_liquido = 0.0
            total_frete_internacional = 0.0

            # Iterar pelas linhas da tabela
            for row_idx in range(self.table.rowCount()):
                descricao = self.table.item(row_idx, 0).text()  # Coluna "Discriminação das despesas não tributável"
                valor_liquido_item = self.table.item(row_idx, 9)  # Coluna "Valor Líquido"

                # Verificar se o item tem um valor válido
                if valor_liquido_item and valor_liquido_item.text() not in ["--", "", None]:
                    try:
                        valor_liquido = float(valor_liquido_item.text().replace(".", "").replace(",", "."))
                    except ValueError:
                        raise ValueError(f"Valor inválido encontrado na coluna 'Valor Líquido': {valor_liquido_item.text()}")

                    # Verificar se é FRETE INTERNACIONAL
                    if descricao and descricao.strip().upper() == "FRETE INTERNACIONAL":
                        total_frete_internacional += valor_liquido
                    else:
                        total_valor_liquido += valor_liquido

            # Calcular os três campos
            campo1 = total_nao_trib - total_frete_internacional
            campo2 = total_valor_liquido
            campo3 = campo1 - campo2

            # Atualizar os campos de Check
            self.field_check_1.setText(f"{campo1:,.2f}".replace(".", ","))
            self.field_check_2.setText(f"{campo2:,.2f}".replace(".", ","))
            self.field_check_3.setText(f"{campo3:,.2f}".replace(".", ","))

        except ValueError as e:
            QMessageBox.critical(self, "Erro", f"Certifique-se de que os valores estão corretos.\n\n{e}")

    def calcular_total_bruto(self):
        try:
            total_bruto = 0.0
            for row_idx in range(self.table.rowCount()):
                valor_bruto_item = self.table.item(row_idx, 11)  # Coluna "VALOR BRUTO" (índice 11)
                if valor_bruto_item:
                    valor_bruto_str = valor_bruto_item.text()
                    if valor_bruto_str not in ["", "--"]:
                        # Corrigindo a conversão para float
                        valor_bruto_str = valor_bruto_str.replace(".", "").replace(",", ".")
                        total_bruto += float(valor_bruto_str)

            # Arredonda o total_bruto para 2 casas decimais
            total_bruto = round(total_bruto, 2)  

            # Exibe o total_bruto no campo total_bruto_input
            self.total_bruto_input.setText(f"{total_bruto:,.2f}".replace(".", ","))

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao calcular o total bruto: {e}")
            
    def calcular_total_bruto_cd(self):
        try:
            total_bruto_cd = 0.0
            for row_idx in range(self.table.rowCount()):
                cfop_item = self.table.item(row_idx, 5)  # Coluna 6 (CFOP)
                if cfop_item and "CD" in cfop_item.text():
                    valor_bruto_item = self.table.item(row_idx, 11)  # Coluna 12 (VALOR BRUTO)
                    if valor_bruto_item:
                        valor_bruto_str = valor_bruto_item.text()
                        if valor_bruto_str not in ["", "--"]:
                            valor_bruto_str = valor_bruto_str.replace(".", "").replace(",", ".")
                            total_bruto_cd += float(valor_bruto_str)

            self.total_cd_input.setText(f"{total_bruto_cd:,.2f}".replace(".", ","))

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao calcular o total bruto CD: {e}")

    def calcular_total_bruto_ca(self):
        try:
            total_bruto_ca = 0.0
            for row_idx in range(self.table.rowCount()):
                cfop_item = self.table.item(row_idx, 5)  # Coluna 6 (CFOP)
                if cfop_item and "CA" in cfop_item.text():
                    valor_bruto_item = self.table.item(row_idx, 11)  # Coluna 12 (VALOR BRUTO)
                    if valor_bruto_item:
                        valor_bruto_str = valor_bruto_item.text()
                        if valor_bruto_str not in ["", "--"]:
                            valor_bruto_str = valor_bruto_str.replace(".", "").replace(",", ".")
                            total_bruto_ca += float(valor_bruto_str)

            self.total_cabruto_input.setText(f"{total_bruto_ca:,.2f}".replace(".", ","))

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao calcular o total bruto CA: {e}")

    def calcular_total_liquido_ca(self):
        try:
            total_liquido_ca = 0.0
            for row_idx in range(self.table.rowCount()):
                cfop_item = self.table.item(row_idx, 5)  # Coluna 6 (CFOP)
                if cfop_item and "CA" in cfop_item.text():
                    valor_liquido_item = self.table.item(row_idx, 9)  # Coluna 10 (VALOR LIQUIDO)
                    if valor_liquido_item:
                        valor_liquido_str = valor_liquido_item.text()
                        if valor_liquido_str not in ["", "--"]:
                            valor_liquido_str = valor_liquido_str.replace(".", "").replace(",", ".")
                            total_liquido_ca += float(valor_liquido_str)

            self.total_caliqui_input.setText(f"{total_liquido_ca:,.2f}".replace(".", ","))

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao calcular o total líquido CA: {e}")
    
    def update_control_excel(self):
        if not self.controle_path:
            QMessageBox.critical(self, "Erro", "Caminho da planilha de controle não foi carregado corretamente.")
            return

        try:
            # Carregar a planilha de controle
            wb = load_workbook(self.controle_path)
            sheet = wb.active

            # Obter a referência da interface gráfica
            referencia = self.reference_input.text()

            # Encontrar a linha correspondente na planilha de controle
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Começar da linha 2
                if row[29] == referencia:  # Comparar com a coluna AD (índice 29)
                    # Iterar sobre as linhas da tabela
                    for table_row_idx in range(self.table.rowCount()):
                        # Obter o valor da coluna 3 (NF FILHA)
                        nf_filha_item = self.table.item(table_row_idx, 2)  # Coluna 3 (índice 2)
                        if nf_filha_item:
                            nf_filha_valor = nf_filha_item.text()

                            # Atualizar a coluna AE (índice 30) com o valor da NF FILHA
                            sheet.cell(row=row_idx, column=31).value = nf_filha_valor
                            break  # Sair do loop interno após encontrar a linha
                    break  # Sair do loop externo após atualizar a planilha de controle

            # Salvar a planilha de controle
            wb.save(self.controle_path)

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar a planilha de controle: {e}")
                


    def export_to_excel(self):
        """
        Exporta os dados de todas as tabelas importadas para um arquivo Excel,
        garantindo que colunas numéricas sejam reconhecidas como números no Excel.
        """
        if not self.tabelas_importadas:
            QMessageBox.warning(self, "Aviso", "Não há dados para exportar!")
            return

        # Abrir diálogo para selecionar o local de salvamento
        file_path, _ = QFileDialog.getSaveFileName(self, "Salvar como Excel", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for idx, (table_name, table_data) in enumerate(self.tabelas_importadas.items(), start=1):
                    if not isinstance(table_data, dict) or "data" not in table_data:
                        QMessageBox.warning(self, "Aviso", f"Dados inválidos para {table_name}. Ignorando...")
                        continue

                    # Extrair os dados
                    referencia = table_data.get("referencia", "Não Informado")
                    importador = table_data.get("importador", "Não Informado")
                    total_nao_trib = table_data.get("total_nao_trib", "Não Informado")
                    tabela_dados = table_data.get("data", [])

                    # Criar DataFrame com os dados
                    df = pd.DataFrame(tabela_dados, columns=HEADERS)

                    # Identificar colunas numéricas e convertê-las para números
                    numeric_columns = ["Total Não Trib.", "Total Bruto", "Total Bruto CD", "Total Bruto CA", "Total Líquido CA"]
                    for col in numeric_columns:
                        if col in df.columns:
                            df[col] = df[col].replace(",", ".", regex=True)  # Substituir vírgula por ponto
                            df[col] = pd.to_numeric(df[col], errors='coerce')  # Converter para float

                    # Exportar para o Excel
                    sheet_name = f"PDF {idx}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)

                    # Obter a planilha
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]

                    # Cores e estilos para cabeçalhos
                    header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
                    header_font = Font(color="ffffff", bold=True)
                    alignment = Alignment(horizontal="center", vertical="center")

                    # Aplicar estilo no cabeçalho (linha 5)
                    for cell in worksheet[5]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = alignment

                    # Adicionar informações complementares (linha 1 a 3)
                    info_rows = {
                        "A1": f"Referência: {referencia}",
                        "A2": f"Importador: {importador}",
                        "A3": f"Total Não Trib.: {total_nao_trib}",
                        "B3": f"Total Bruto: {self.total_bruto_input.text()}",
                        "C3": f"Total Bruto CD: {self.total_cd_input.text()}",
                        "D3": f"Total Bruto CA: {self.total_cabruto_input.text()}",
                        "E3": f"Total Líquido CA: {self.total_caliqui_input.text()}",
                    }

                    for cell, value in info_rows.items():
                        worksheet[cell] = value
                        worksheet[cell].fill = header_fill
                        worksheet[cell].font = header_font
                        worksheet[cell].alignment = alignment

                    # Formatar as colunas B, J e L como numéricas
                    for row in worksheet.iter_rows(min_row=5):  # Começa na linha 5
                        for cell in row:
                            if cell.column_letter in ['B', 'J', 'L', 'N', 'O', 'P']:  # Verifica se a célula está nas colunas B, J ou L
                                try:
                                    # Tentar converter o valor da célula para numérico
                                    valor = float(str(cell.value).replace('.', '').replace(',', '.'))
                                    cell.value = valor
                                    cell.number_format = '#,##0.00'  # Formato numérico com duas casas decimais
                                except ValueError:
                                    pass  # Ignorar células que não podem ser convertidas


                    # Ajustar largura das colunas
                    for column_cells in worksheet.columns:
                        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
                        col_letter = column_cells[0].column_letter
                        worksheet.column_dimensions[col_letter].width = max_length + 2

                    # Remover linhas de grade
                    worksheet.sheet_view.showGridLines = False

            # Perguntar ao usuário se deseja atualizar a planilha de controle
            reply = QMessageBox.question(
                self, "Atualizar Planilha de Controle",
                "Deseja atualizar a planilha de controle?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                try:
                    self.update_control()
                    QMessageBox.information(self, "Sucesso", "Planilha de controle atualizada com sucesso!")
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao atualizar a planilha de controle: {e}")

            QMessageBox.information(self, "Sucesso", f"Dados exportados para {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar os dados: {e}")








if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PDFTableExtractorApp()
    window.show()
    sys.exit(app.exec())
